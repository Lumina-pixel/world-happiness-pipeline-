import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Load and clean the Happiness Index data
df = pd.read_csv('WHR2023.csv')

# Fill missing values with median
df['Healthy life expectancy'] = df['Healthy life expectancy'].fillna(df['Healthy life expectancy'].median())
df['Explained by: Healthy life expectancy'] = df['Explained by: Healthy life expectancy'].fillna(df['Explained by: Healthy life expectancy'].median())
df['Dystopia + residual'] = df['Dystopia + residual'].fillna(df['Dystopia + residual'].median())

# Rename columns to match hapiness.py renamed columns (and clean others where helpful)
rename_dict = {
    'Country name': 'Country',
    'Ladder score': 'Happiness_Score',
    'Logged GDP per capita': 'GDP',
    'Social support': 'Social_Support',
    'Healthy life expectancy': 'Life_Expectancy',
    'Freedom to make life choices': 'Freedom',
    'Perceptions of corruption': 'Corruption'
}
df_renamed = df.rename(columns=rename_dict)

# Calculate statistics for the summary card
avg_happiness = df_renamed['Happiness_Score'].mean()
median_happiness = df_renamed['Happiness_Score'].median()
total_countries = len(df_renamed)
happiest_country = df_renamed.loc[df_renamed['Happiness_Score'].idxmax()]['Country']
happiest_score = df_renamed['Happiness_Score'].max()
least_happy_country = df_renamed.loc[df_renamed['Happiness_Score'].idxmin()]['Country']
least_happy_score = df_renamed['Happiness_Score'].min()

# 2. Extract analytical tables as specified in hapiness.py
# A. Top 10 Most Corrupt Countries
top_corrupt = df_renamed[['Country', 'Corruption']].sort_values(by='Corruption', ascending=False).head(10)

# B. Top 10 GDP Countries (sorted by Happiness Score)
top_gdp = df_renamed.sort_values(by='GDP', ascending=False).head(10)
top_gdp_sorted = top_gdp[['Country', 'GDP', 'Happiness_Score']].sort_values(by='Happiness_Score')

# C. Top 50 Free Countries (sorted by Happiness Score)
top_free = df_renamed.sort_values(by='Freedom', ascending=False).head(50)
top_free_sorted = top_free[['Country', 'Freedom', 'Happiness_Score']].sort_values(by='Happiness_Score')

# D. India vs China vs Pakistan comparison
incnpk = df_renamed[df_renamed['Country'].isin(['India', 'China', 'Pakistan'])][
    ['Country', 'Happiness_Score', 'GDP', 'Freedom', 'Corruption']
].copy()
incnpk['Greater_than_average_happiness'] = incnpk['Happiness_Score'].map(lambda x: 'True' if x > median_happiness else 'False')

# 3. Create Workbook
wb = openpyxl.Workbook()

# Define Color Palette (Teal Theme)
header_fill = PatternFill(start_color="005B5C", end_color="005B5C", fill_type="solid") # Dark Teal
zebra_fill = PatternFill(start_color="F2FAFA", end_color="F2FAFA", fill_type="solid") # Very Light Teal
accent_fill = PatternFill(start_color="D6ECEC", end_color="D6ECEC", fill_type="solid") # Light Teal Accent
title_fill = PatternFill(start_color="003C3D", end_color="003C3D", fill_type="solid") # Deep Navy Teal

font_family = "Segoe UI"
title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
section_font = Font(name=font_family, size=13, bold=True, color="003C3D")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
bold_font = Font(name=font_family, size=11, bold=True, color="000000")
regular_font = Font(name=font_family, size=11, color="000000")
italic_font = Font(name=font_family, size=9, italic=True, color="555555")

thin_border_side = Side(style='thin', color='DDDDDD')
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
thick_bottom = Border(bottom=Side(style='medium', color='003C3D'))
double_bottom = Border(top=Side(style='thin', color='DDDDDD'), bottom=Side(style='double', color='003C3D'))

align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')

# ----------------- SHEET 1: OVERVIEW -----------------
ws_info = wb.active
ws_info.title = "Overview & Summary"
ws_info.views.sheetView[0].showGridLines = True

# Title Block
ws_info.merge_cells('A1:F2')
for r in range(1, 3):
    for c in range(1, 7):
        cell = ws_info.cell(row=r, column=c)
        cell.fill = title_fill
title_cell = ws_info.cell(row=1, column=1)
title_cell.value = "World Happiness Report 2023 Overview"
title_cell.font = title_font
title_cell.alignment = align_center

# Description
ws_info.cell(row=4, column=1, value="About this dataset:").font = section_font
ws_info.cell(row=5, column=1, value="The World Happiness Report is a landmark survey of the state of global happiness.").font = regular_font
ws_info.cell(row=6, column=1, value="It ranks countries by their happiness levels based on life evaluations (Ladder Score).").font = regular_font
ws_info.cell(row=7, column=1, value="This Excel workbook provides structured views of the happiness dataset for 2023, along with critical comparisons.").font = regular_font

# Summary Cards Header
ws_info.cell(row=9, column=1, value="Key Performance Indicators (KPIs)").font = section_font

# KPI Card 1
ws_info.merge_cells('B11:C11')
ws_info.merge_cells('B12:C12')
card1_label = ws_info.cell(row=11, column=2, value="TOTAL COUNTRIES")
card1_label.font = Font(name=font_family, size=9, bold=True, color="555555")
card1_label.alignment = align_center
card1_val = ws_info.cell(row=12, column=2, value=total_countries)
card1_val.font = Font(name=font_family, size=18, bold=True, color="005B5C")
card1_val.alignment = align_center
for r in range(11, 13):
    for c in range(2, 4):
        ws_info.cell(row=r, column=c).fill = zebra_fill
        ws_info.cell(row=r, column=c).border = thin_border

# KPI Card 2
ws_info.merge_cells('E11:F11')
ws_info.merge_cells('E12:F12')
card2_label = ws_info.cell(row=11, column=5, value="AVERAGE HAPPINESS SCORE")
card2_label.font = Font(name=font_family, size=9, bold=True, color="555555")
card2_label.alignment = align_center
card2_val = ws_info.cell(row=12, column=5, value=avg_happiness)
card2_val.font = Font(name=font_family, size=18, bold=True, color="005B5C")
card2_val.alignment = align_center
card2_val.number_format = '0.000'
for r in range(11, 13):
    for c in range(5, 7):
        ws_info.cell(row=r, column=c).fill = zebra_fill
        ws_info.cell(row=r, column=c).border = thin_border

# KPI Card 3
ws_info.merge_cells('B14:C14')
ws_info.merge_cells('B15:C15')
card3_label = ws_info.cell(row=14, column=2, value="HAPPIEST COUNTRY")
card3_label.font = Font(name=font_family, size=9, bold=True, color="555555")
card3_label.alignment = align_center
card3_val = ws_info.cell(row=15, column=2, value=f"{happiest_country} ({happiest_score:.3f})")
card3_val.font = Font(name=font_family, size=11, bold=True, color="005B5C")
card3_val.alignment = align_center
for r in range(14, 16):
    for c in range(2, 4):
        ws_info.cell(row=r, column=c).fill = zebra_fill
        ws_info.cell(row=r, column=c).border = thin_border

# KPI Card 4
ws_info.merge_cells('E14:F14')
ws_info.merge_cells('E15:F15')
card4_label = ws_info.cell(row=14, column=5, value="LEAST HAPPY COUNTRY")
card4_label.font = Font(name=font_family, size=9, bold=True, color="555555")
card4_label.alignment = align_center
card4_val = ws_info.cell(row=15, column=5, value=f"{least_happy_country} ({least_happy_score:.3f})")
card4_val.font = Font(name=font_family, size=11, bold=True, color="005B5C")
card4_val.alignment = align_center
for r in range(14, 16):
    for c in range(5, 7):
        ws_info.cell(row=r, column=c).fill = zebra_fill
        ws_info.cell(row=r, column=c).border = thin_border

# Column Legend Table
ws_info.cell(row=18, column=1, value="Data Dictionary & Columns Guide").font = section_font

legend_headers = ["Column Name", "Description"]
for c_idx, h in enumerate(legend_headers, start=1):
    cell = ws_info.cell(row=20, column=c_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_left
    cell.border = thin_border

legend_data = [
    ("Country", "Name of the country"),
    ("Happiness_Score", "Ladder score representing the self-reported happiness level (0 to 10)"),
    ("GDP", "Logged GDP per capita"),
    ("Social_Support", "National average of the binary responses (0 or 1) to the social support question"),
    ("Life_Expectancy", "Healthy life expectancy at birth"),
    ("Freedom", "National average of responses to the freedom to make life choices question"),
    ("Corruption", "National average of responses to perceptions of corruption in government/business"),
    ("Dystopia + residual", "Calculated value comparing each country to a hypothetical worst-case country (Dystopia)")
]

for r_idx, (col_n, col_d) in enumerate(legend_data, start=21):
    c1 = ws_info.cell(row=r_idx, column=1, value=col_n)
    c2 = ws_info.cell(row=r_idx, column=2, value=col_d)
    c1.font = bold_font
    c2.font = regular_font
    c1.border = thin_border
    c2.border = thin_border
    if r_idx % 2 == 0:
        c1.fill = zebra_fill
        c2.fill = zebra_fill


# ----------------- SHEET 2: DATA -----------------
ws_data = wb.create_sheet(title="Happiness Data")
ws_data.views.sheetView[0].showGridLines = True

# Write DataFrame headers
for col_idx, column in enumerate(df_renamed.columns, start=1):
    cell = ws_data.cell(row=1, column=col_idx, value=column)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border

# Write DataFrame values
for row_idx, row in enumerate(df_renamed.values, start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
        cell.font = regular_font
        cell.border = thin_border
        
        # Zebra striping
        if row_idx % 2 == 1:
            cell.fill = zebra_fill
            
        # Format alignment and numbers
        if isinstance(val, (int, float)):
            cell.alignment = align_right
            if col_idx == 8: # Healthy life expectancy has values like 71.15
                cell.number_format = '0.0'
            else:
                cell.number_format = '0.000'
        else:
            cell.alignment = align_left


# ----------------- SHEET 3: INSIGHTS -----------------
ws_insights = wb.create_sheet(title="Analysis & Insights")
ws_insights.views.sheetView[0].showGridLines = True

# Helper function to write tables with styling
def write_styled_table(ws, start_row, start_col, title, df_table, col_formats=None):
    # Section Title
    cell_title = ws.cell(row=start_row, column=start_col, value=title)
    cell_title.font = section_font
    
    # Headers
    header_row = start_row + 1
    for col_idx, col_name in enumerate(df_table.columns):
        cell = ws.cell(row=header_row, column=start_col + col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Data Rows
    current_row = header_row + 1
    for i, row in enumerate(df_table.values):
        for col_idx, val in enumerate(row):
            cell = ws.cell(row=current_row + i, column=start_col + col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            
            # Format alignment and numbers
            if isinstance(val, (int, float)):
                cell.alignment = align_right
                if col_formats and col_idx in col_formats:
                    cell.number_format = col_formats[col_idx]
                else:
                    cell.number_format = '0.000'
            else:
                cell.alignment = align_left
                
            # Zebra striping
            if (current_row + i) % 2 == 1:
                cell.fill = zebra_fill
                
    return current_row + len(df_table)

# 1. India, China, Pakistan Comparison
next_r = write_styled_table(
    ws_insights, 
    start_row=2, 
    start_col=1, 
    title="India vs China vs Pakistan Comparison", 
    df_table=incnpk,
    col_formats={1: '0.000', 2: '0.000', 3: '0.000', 4: '0.000'}
)

# Highlight Comparison Row
for c in range(1, 7):
    ws_insights.cell(row=next_r-3, column=c).fill = accent_fill # Highlight India (usually the user's primary region)
    ws_insights.cell(row=next_r-3, column=c).font = bold_font

# 2. Top 10 Most Corrupt Countries
next_r_corrupt = write_styled_table(
    ws_insights, 
    start_row=next_r + 3, 
    start_col=1, 
    title="Top 10 Countries with Highest Perceived Corruption", 
    df_table=top_corrupt,
    col_formats={1: '0.000'}
)

# 3. Top 10 GDP Countries (Sorted by Happiness)
next_r_gdp = write_styled_table(
    ws_insights, 
    start_row=2, 
    start_col=8, 
    title="Top 10 GDP Countries (Sorted by Happiness Score)", 
    df_table=top_gdp_sorted,
    col_formats={1: '0.000', 2: '0.000'}
)

# 4. Top 50 Free Countries (Sorted by Happiness Score)
next_r_free = write_styled_table(
    ws_insights, 
    start_row=next_r_gdp + 3, 
    start_col=8, 
    title="Top 50 Free Countries (Sorted by Happiness Score)", 
    df_table=top_free_sorted,
    col_formats={1: '0.000', 2: '0.000'}
)


# ----------------- AUTO-FIT COLUMN WIDTHS FOR ALL SHEETS -----------------
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't auto-fit merged cells or very long strings on overview sheet to prevent huge column widths
        for cell in col:
            # Skip checking length if it is merged or a long description
            if ws.title == "Overview & Summary" and cell.row <= 7:
                continue
            if cell.value:
                # Add padding for format characters
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        
        # Add safety margin
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save the beautifully styled Excel workbook
wb.save('WHR2023.xlsx')
print("Successfully generated WHR2023.xlsx!")
